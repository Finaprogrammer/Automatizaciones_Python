import openpyxl
from openpyxl.styles import Font, PatternFill, alignment
import os  # Importar el módulo os
def crear_excel_inicial():
    #crea el archivo
    wb = openpyxl.Workbook() #crea el archivo de un workbook
    #crea la hoja
    ws = wb.active
    ws.title = "Inventario principal"
    
    #definir los encabezados
    encabezados = ["ID", "producto", "categoria", "precio", "stock"]
    # Lista de listas con productos tecnológicos
    productos = [
        [1, "HP Pavilion", "PC", 1000, 10],
        [2, "Ipad", "Tablet", 500, 5],
        [3, "Iphone 13", "Smartphone", 300, 30],
        [4, "Apple watch", "Smartwatch", 200, 40],
        [5, "Beats", "Auriculares", 100, 50]
    ]

    # Iterar sobre los encabezados y agregarlos a la hoja
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.font = Font(bold=True)
        celda.fill = PatternFill(start_color="CCCCCC", end_color="FFD700", fill_type="solid")
        celda.alignment = alignment.Alignment(horizontal="center")
    
    # escribir los productos

    for fila, producto in enumerate(productos, start=2):
        for col, valor in enumerate(producto, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = alignment.Alignment(horizontal="center")
    #ajustar anchos de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Obtener la ruta del archivo actual y guardar el Excel en la misma carpeta

    ruta_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_archivo = os.path.join(ruta_actual, "inventario_tecnologia.xlsx")
    wb.save(ruta_archivo)

    

if __name__ == '__main__':
    crear_excel_inicial()
    print("Archivo creado")
