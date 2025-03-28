import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart
import os
from openpyxl.worksheet.table import Table, TableStyleInfo

def crear_grafico_ventas(ws):
    """Crear un gráfico de barras con el stock de los productos."""
    # Crear gráfico de barras
    chart = BarChart()
    chart.title = "Stock de Productos"
    chart.y_axis.title = 'Cantidad'
    chart.x_axis.title = 'Productos'

    # Obtener los datos para el gráfico
    data = Reference(
        worksheet=ws,
        min_row=2,  # Start from row 2 to skip header
        max_row=ws.max_row,
        min_col=5,  # Stock column
        max_col=5
    )

    # Obtener las categorías (nombres de productos)
    cats = Reference(
        worksheet=ws,
        min_row=2,  # Start from row 2 to skip header
        max_row=ws.max_row,
        min_col=2,  # Product name column
        max_col=2
    )

    # Agregar los datos al gráfico
    chart.add_data(data)
    chart.set_categories(cats)

    # Ajustar el tamaño del gráfico
    chart.height = 10
    chart.width = 15

    # Agregar el gráfico a la hoja
    ws.add_chart(chart, "G2")

def crear_grafico_categoria(ws):
    """crear grafico circular para visualizar distribucion por categorias"""
    grafico=PieChart()
    grafico.title="Distribucion por categorias"
    #datos para el grafico
    data=Reference(ws, min_col=5, min_row=1, max_row=ws.max_row, max_col=5)
    etiquetas=Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    grafico.add_data(data, titles_from_data=True)
    grafico.set_categories(etiquetas)
    #añadir grafico a la hoja
    ws.add_chart(grafico, "G18")

def crear_tabla(wb, ws):
    """"crear tabla"""
    #convertir rango de datos en tabla
    tab = Table(displayName="Inventario", ref=f"A1:E{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    




def automatizacion_avanzada():
    """Función principal para la automatización avanzada."""
    try:
        # Cargar el archivo existente
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        ruta_archivo = os.path.join(ruta_actual, "inventario_tecnologia.xlsx")
        
        # Cargar el workbook con data_only=True para obtener valores calculados
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        ws = wb.active

        print("Creando gráfico de ventas...")
        crear_grafico_ventas(ws)
        crear_grafico_categoria(ws)

        #crear tabla
        print("Creando tabla...")
        crear_tabla(wb, ws)

        # Guardar el archivo
        ruta_salida = os.path.join(ruta_actual, "inventario_tecnologia_v2.xlsx")
        wb.save(ruta_salida)
        print(f"Archivo guardado exitosamente como: inventario_tecnologia_con_grafico.xlsx")
        
    except FileNotFoundError:
        print("Error: No se encontró el archivo inventario_tecnologia.xlsx")
    except Exception as e:
        print(f"Error inesperado: {str(e)}")

if __name__ == '__main__':
    automatizacion_avanzada()