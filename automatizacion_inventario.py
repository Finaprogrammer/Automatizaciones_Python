import openpyxl
from openpyxl.styles import Font, PatternFill, alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def cargar_inventario():
    try:
        # Obtener la ruta del archivo actual
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        ruta_archivo = os.path.join(ruta_actual, "inventario_tecnologia.xlsx")
        
        # Cargar el archivo de inventario
        wb = openpyxl.load_workbook(ruta_archivo)
    except FileNotFoundError:
        print("No se encontró el archivo de inventario en la carpeta actual.")
        return None, None
    
    ws = wb.active
    ws.title = "Inventario principal"

    return wb, ws
def actualizar_precios(ws, porcentaje):
    """Actualizar los precios de los productos en el inventario segun inventario"""
    for row in range(2, ws.max_row + 1):
        precio_actual = float(ws.cell(row=row, column=4).value)
        nuevo_precio = precio_actual * (1 + porcentaje/100)
        ws.cell(row=row, column=4, value= round(nuevo_precio, 2))

def verificar_stock_bajo(ws, cantidad_minima=15):
    """Verificar si hay productos con stock bajo en el inventario."""
    productos_bajos_stock = []
    for row in range(2, ws.max_row + 1):
        stock = ws.cell(row=row, column=5).value
        if stock < cantidad_minima:
            producto = ws.cell(row=row, column=2).value
            productos_bajos_stock.append((producto, stock))
            # Marcar el producto con color rojo
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    return productos_bajos_stock

def generar_reporte():
    """Generar un reporte con los productos con stock bajo."""
    wb, ws = cargar_inventario()
    if not wb:
        return
    
    #crear una nueva hoja para el reporte

    fecha_actual = datetime.now().strftime("%d-%m-%Y")
    ws_reporte = wb.create_sheet(title=f"Reporte {fecha_actual}")

    #copiar los encabezados
    for col in range(1, ws.max_column + 1):
       ws_reporte.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)
       ws_reporte.cell(row=1, column=col).font = Font(bold=True)
    #copiar datos
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws_reporte.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)
    # añadir estadisticas
    row_estadisticas = ws.max_row + 2
    ws_reporte.cell(row=row_estadisticas, column=1, value="estadisticas del inventario")
    ws_reporte.cell(row=row_estadisticas+1, column=1, value="productos diferentes:")
    ws_reporte.cell(row=row_estadisticas+1, column=2, value=f"=COUNTA(E2:E{ws.max_row})")
    ws_reporte.cell(row=row_estadisticas+2, column=1, value="total unidades:")
    ws_reporte.cell(row=row_estadisticas+2, column=2, value=f"=SUM(E2:E{ws.max_row})")
    ws_reporte.cell(row=row_estadisticas+3, column=1, value="valor total:")
    ws_reporte.cell(row=row_estadisticas+3, column=2, value=f"=SUM(D2:D{ws.max_row})")
    #ajustar anchos de las columnas
    for col in ws_reporte.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_reporte.column_dimensions[column].width = adjusted_width 


    return wb



def automatizacion_inventario():
    """Automatizar la creación de un archivo Excel con productos tecnológicos."""
    wb, ws = cargar_inventario()
    if not wb:
        return
    print("\nsistema de automatización de inventario")
    actualizar_precios(ws, porcentaje=5)
    print("\nActualización de precios con un incremento del 5%.")
    # Verificar productos con stock bajo
    productos_bajos_stock = verificar_stock_bajo(ws)
    if productos_bajos_stock:
        print("\nProductos con stock bajo:")
        for producto, stock in productos_bajos_stock:
            print(f"{producto}: {stock} unidades.")
    wb = generar_reporte()

    # Guardar el archivo actualizado
    ruta_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_archivo_v2 = os.path.join(ruta_actual, "inventario_tecnologia_v2.xlsx")
    wb.save(ruta_archivo_v2)
    print("\nProceso de automatización finalizado. Archivo guardado como 'inventario_tecnologia_v2.xlsx'.")

if __name__ == '__main__':
    automatizacion_inventario()






